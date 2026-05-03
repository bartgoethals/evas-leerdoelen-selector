from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "bronnen"

DISCIPLINE_HINTS = [
    ("ned_com", "Nederlands en communicatie"),
    ("w_t", "Wetenschap en techniek"),
    ("wiskunde", "Wiskunde"),
    ("frans", "Frans"),
    ("ict", "ICT"),
    ("lele", "Leren leren"),
    ("muvo", "MUVO"),
    ("aardr", "Aardrijkskunde"),
    ("aardrijkskunde", "Aardrijkskunde"),
    ("gesch", "Geschiedenis"),
    ("veilig_gezond", "Veilig en gezond"),
    ("v_g", "Veilig en gezond"),
    ("rkg", "Rooms-katholieke godsdienst"),
    ("godsdienst", "Rooms-katholieke godsdienst"),
    ("lo_", "Lichamelijke opvoeding"),
    ("-lo", "Lichamelijke opvoeding"),
    ("_lo", "Lichamelijke opvoeding"),
]

WORKBOOKS = sorted(SOURCE_DIR.glob("*.xlsx"))
PDF_FILES = sorted(SOURCE_DIR.glob("*.pdf"))


def detect_discipline(name: str, default: str = "Onbekend") -> str:
    lname = name.lower()
    for token, label in DISCIPLINE_HINTS:
        if token in lname:
            return label
    return default


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def safe_token(value: str) -> str:
    token = (
        value.strip()
        .lower()
        .replace(" ", "-")
        .replace("/", "-")
        .replace("\\", "-")
        .replace("_", "-")
        .replace(".", "-")
    )
    while "--" in token:
        token = token.replace("--", "-")
    return token.strip("-")


def normalize_header(value: str) -> str:
    return (
        value.lower()
        .replace("-", "")
        .replace("_", "")
        .replace(" ", "")
        .replace("/", "")
        .replace("(", "")
        .replace(")", "")
        .replace(".", "")
        .replace("&", "en")
    )


def val(ws, row: int, header_idx: dict, *names: str, fallback_col: int | None = None) -> str:
    for name in names:
        col = header_idx.get(normalize_header(name))
        if col:
            return clean(ws.cell(row=row, column=col).value)
    if fallback_col:
        return clean(ws.cell(row=row, column=fallback_col).value)
    return ""


def load_goals() -> list[dict]:
    results: list[dict] = []
    for workbook_path in WORKBOOKS:
        subject = detect_discipline(workbook_path.stem, workbook_path.stem)
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [clean(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
        header_idx = {normalize_header(h): i + 1 for i, h in enumerate(headers) if h}

        for row in range(2, ws.max_row + 1):
            goal_text = val(ws, row, header_idx, "Leerplandoel")
            if not goal_text:
                continue

            code = val(ws, row, header_idx, "Code")
            unique_key = safe_token(code) if code else f"r{row}"
            wb_slug = safe_token(workbook_path.stem)
            item = {
                "id": f"{wb_slug}-{unique_key}",
                "vak": subject,
                "doelsoort": val(ws, row, header_idx, "Doel-soort", "Doel soort", "Doelsoort"),
                "lfmd": val(ws, row, header_idx, "LfMD", fallback_col=2),
                "nrmd": val(ws, row, header_idx, "NrMD", "nrMD", "Nr MD", "nr MD", fallback_col=3),
                "md": val(ws, row, header_idx, "MD", fallback_col=4),
                "code": code,
                "fase": val(ws, row, header_idx, "Jaar/ fase", "Jaar/\nfase"),
                "domein": val(ws, row, header_idx, "Domein"),
                "subdomein": val(ws, row, header_idx, "Subdomein"),
                "cluster": val(ws, row, header_idx, "Cluster"),
                "leerplandoel": goal_text,
                "voorbeelden": val(ws, row, header_idx, "Voorbeelden", "Voorbeeld"),
                "extra_toelichting": val(ws, row, header_idx, "Extra toelichting", "Toelichting", "Korte toelichting"),
                "woordenschat": val(ws, row, header_idx, "Woordenschat (richtinggevend)"),
            }
            results.append(item)

    return results


def build_resources() -> list[dict]:
    resources: list[dict] = []
    for path in PDF_FILES:
        filename = path.name
        subject = detect_discipline(path.stem)
        size_kb = round(path.stat().st_size / 1024)
        resources.append(
            {
                "vak": subject,
                "bestand": filename,
                "titel": filename.replace("_", " ").replace(".pdf", ""),
                "url": f"bronnen/{filename}",
                "grootte_kb": size_kb,
            }
        )
    return resources


def main() -> None:
    if not SOURCE_DIR.exists():
        raise SystemExit("Bronmap ontbreekt: ./bronnen")
    goals = load_goals()
    payload = {
        "meta": {
            "bron": "Nieuwe leerdoelen (Excel + PDF) uit /bronnen",
            "aantal": len(goals),
            "vakken": sorted({g["vak"] for g in goals}),
        },
        "doelen": goals,
        "bronnen": build_resources(),
    }
    out = ROOT / "data" / "goals.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out} with {len(goals)} doelen")


if __name__ == "__main__":
    main()
